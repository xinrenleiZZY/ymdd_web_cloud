import streamlit as st
import pandas as pd
import os
import sys
from datetime import datetime
import traceback
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
from openpyxl.styles import NamedStyle
import tempfile
from io import BytesIO, StringIO
import requests

# 新增：配置GitHub仓库信息（请替换为你的实际信息）
GITHUB_REPO_INFO = {
    "username": "xinrenleiZZY",
    "repo_name": "ymdd_web_cloud",
    "branch": "master",  # master
    "hidden_file_path": "mnt/隐藏表格.xlsx"  # mnt文件夹下的隐藏表格路径
}

def print_banner():
    """显示程序标题"""
    st.markdown("""
    ---
    ### 益模订单转换工具 v1.0
    **非开发人员专用版本**
    ---
    """)

def copy_sheet(source_wb, source_sheet_name, target_wb, new_sheet_name=None):
    """复制工作表（包含完整格式）"""
    source_sheet = source_wb[source_sheet_name]
    new_name = new_sheet_name or source_sheet_name
    target_sheet = target_wb.create_sheet(new_name)

    # 复制单元格内容和样式
    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row,
                                     min_col=1, max_col=source_sheet.max_column):
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            
    # 复制列宽
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, source_sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    
    # 复制行高
    for row in range(1, source_sheet.max_row + 1):
        if row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    # 复制合并单元格
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merged_cells.add(str(merged_range))

    # 复制工作表属性
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.conditional_formatting = copy(source_sheet.conditional_formatting)

    # 修正：正确复制命名样式
    target_style_names = []
    for s in target_wb.named_styles:
        if hasattr(s, 'name'):
            target_style_names.append(s.name)
        elif isinstance(s, str):
            target_style_names.append(s)
    
    for style in source_wb.named_styles:
        if hasattr(style, 'name'):
            style_name = style.name
        elif isinstance(style, str):
            style_name = style
        else:
            continue
            
        if style_name not in target_style_names:
            new_style = NamedStyle(name=style_name)
            if hasattr(style, 'font'):
                new_style.font = copy(style.font)
            if hasattr(style, 'border'):
                new_style.border = copy(style.border)
            if hasattr(style, 'fill'):
                new_style.fill = copy(style.fill)
            if hasattr(style, 'number_format'):
                new_style.number_format = copy(style.number_format)
            if hasattr(style, 'protection'):
                new_style.protection = copy(style.protection)
            if hasattr(style, 'alignment'):
                new_style.alignment = copy(style.alignment)
            target_wb.add_named_style(new_style)

    return target_sheet

# 新增：从GitHub获取隐藏表格
def get_hidden_file_from_github():
    """从GitHub仓库的mnt文件夹读取隐藏表格"""
    try:
        # 构建GitHub原始文件链接
        # https://github.com/xinrenleiZZY/ymdd_web_cloud/blob/master/mnt/%E9%9A%90%E8%97%8F%E8%A1%A8%E6%A0%BC.xlsx
        # 构建正确的GitHub原始文件链接（用于直接下载文件）
        github_url = (
            f"https://raw.githubusercontent.com/"
            f"{GITHUB_REPO_INFO['username']}/"
            f"{GITHUB_REPO_INFO['repo_name']}/"
            f"{GITHUB_REPO_INFO['branch']}/"
            f"{GITHUB_REPO_INFO['hidden_file_path']}"
        )
        
        st.info(f"正在从GitHub获取隐藏表格：{github_url}")
        response = requests.get(github_url)
        response.raise_for_status()  # 若请求失败（如404），抛出异常
        
        # 将请求内容转为文件对象供openpyxl读取
        file_stream = BytesIO(response.content)
        return file_stream
    
    except Exception as e:
        st.error(f"获取隐藏表格失败：{str(e)}")
        st.text("请检查GitHub仓库信息是否正确，或文件路径是否存在")
        return None
    
def convert_files(source_file, hidden_file):
    """执行文件转换并返回结果"""
    try:
        st.info("正在读取何氏订单总表...")
        df_source = pd.read_excel(source_file)
        st.success(f"源文件读取成功，共 {len(df_source)} 行数据")

        # 生成订单录入文件
        st.info("正在生成订单录入文件...")
        df_unique = df_source.drop_duplicates(subset=['生产单号'], keep='first')
        st.success(f"按生产单号去重完成，共 {len(df_unique)} 条记录")

        order_data = []
        for index, row in df_unique.iterrows():
            new_row = {
                '项目名称': str(row['制品名称']),
                '项目编号': str(row['制品名称']),
                '项目预估交货期': row['下单日期'].strftime('%Y-%m-%d'),
                '模具名称': str(row['制品名称']),
                '模具编号': str(row['生产单号']),
                '预估交货期': row['交期'].strftime('%Y-%m-%d'),
                '模具类型': str(row['类型']),
                '模具阶段': str(row['Unnamed: 7']),
                '数量': 1
            }
            order_data.append(new_row)

        df_order_result = pd.DataFrame(order_data)

        # 生成工件导入文件
        st.info("正在生成工件导入文件...")
        workpiece_data = []
        for index, row in df_source.iterrows():
            base_row = {
                '生产任务号': str(row['生产单号']) + '_T0',
                '件号': str(row['制品名称']) + str(row['部件名称']),
                '工件编码': str(row['制品名称']),
                '工件名称': str(row['部件名称']),
                '数量': int(row['数量']),
                '备注': '',
                '生产单号': str(row['生产单号'])
            }
            workpiece_data.append(base_row)

            # 处理母型合金列
            if pd.notna(row.get('母型合金')) and str(row['母型合金']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '母型合金',
                    '工件编码': '母型合金',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理母型合金板列
            if pd.notna(row.get('母型合金板')) and str(row['母型合金板']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '母型合金板',
                    '工件编码': '母型合金板',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理母型套中套列
            if pd.notna(row.get('母型套中套')) and str(row['母型套中套']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '母型套中套',
                    '工件编码': '母型套中套',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理合金针列
            if pd.notna(row.get('合金针')) and str(row['合金针']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '合金针',
                    '工件编码': '合金针',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理底座列
            if pd.notna(row.get('底座')) and str(row['底座']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': f"{row['部件名称']}底座",
                    '工件编码': f"{row['部件名称']}底座",
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

        df_workpiece_result = pd.DataFrame(workpiece_data)
        st.success(f"工件导入数据生成完成，共 {len(df_workpiece_result)} 条记录")

        # 处理隐藏表格文件
        hidden_wb = load_workbook(hidden_file, data_only=True)

        # 生成订单录入文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        order_wb = Workbook()
        if 'Sheet' in order_wb.sheetnames:
            del order_wb['Sheet']
        
        copy_sheet(hidden_wb, 'page', order_wb, new_sheet_name='page')
        order_ws = order_wb.create_sheet('订单录入')
        for r in dataframe_to_rows(df_order_result, index=False, header=True):
            order_ws.append(r)
        
        column_widths = {
            'A': 35, 'B': 35, 'C': 15, 'D': 35, 'E': 12,
            'F': 15, 'G': 20, 'H': 12, 'I': 8
        }
        for col_letter, width in column_widths.items():
            order_ws.column_dimensions[col_letter].width = width
        order_ws.row_dimensions[1].height = 25
        for row_num in range(2, len(df_order_result) + 2):
            order_ws.row_dimensions[row_num].height = 20
        
        order_wb['page'].sheet_state = 'hidden'
        
        # 保存到内存
        order_buffer = BytesIO()
        order_wb.save(order_buffer)
        order_buffer.seek(0)

        # 生成工件导入文件
        workpiece_wb = Workbook()
        if 'Sheet' in workpiece_wb.sheetnames:
            del workpiece_wb['Sheet']
        
        copy_sheet(hidden_wb, 'page2', workpiece_wb, new_sheet_name='page')
        workpiece_ws = workpiece_wb.create_sheet('工件信息')
        for r in dataframe_to_rows(df_workpiece_result, index=False, header=True):
            workpiece_ws.append(r)
        
        workpiece_column_widths = {
            'A': 15, 'B': 50, 'C': 35, 'D': 20, 'E': 8, 'F': 10, 'G': 12
        }
        for col_letter, width in workpiece_column_widths.items():
            workpiece_ws.column_dimensions[col_letter].width = width
        workpiece_ws.row_dimensions[1].height = 25
        for row_num in range(2, len(df_workpiece_result) + 2):
            workpiece_ws.row_dimensions[row_num].height = 20
        
        workpiece_wb['page'].sheet_state = 'hidden'
        
        # 保存到内存
        workpiece_buffer = BytesIO()
        workpiece_wb.save(workpiece_buffer)
        workpiece_buffer.seek(0)

        st.success("转换完成！")
        return {
            'order': {
                'buffer': order_buffer,
                'filename': f'订单录入结果_{timestamp}.xlsx',
                'count': len(df_order_result)
            },
            'workpiece': {
                'buffer': workpiece_buffer,
                'filename': f'工件导入结果_{timestamp}.xlsx',
                'count': len(df_workpiece_result)
            }
        }

    except Exception as e:
        st.error(f"转换过程中出现错误: {str(e)}")
        st.text("详细错误信息:")
        st.text(traceback.format_exc())
        return None


def main():
    """主函数"""
    print_banner()

    st.info("""
    **程序说明：**
    本工具将何氏订单总表的数据转换为两个文件：
    1. 订单录入结果 - 按生产单号去重后的模具级别信息
    2. 工件导入结果 - 包含所有工件详细信息
    """)

    # 上传源文件
    source_file = st.file_uploader("选择何氏订单总表文件（Excel格式）", type=["xlsx"])
    
    # # 上传隐藏表格文件
    # hidden_file = st.file_uploader("选择隐藏表格文件（Excel格式）", type=["xlsx"])

    if st.button("开始转换"):
        if not source_file:
            st.error("请先选择订单总表文件")
            return
        
        # 自动从GitHub获取隐藏表格
        hidden_file = get_hidden_file_from_github()
        if not hidden_file:
            st.error("无法获取隐藏表格，转换终止")
            return
        # if not hidden_file:
        #     st.error("请先选择隐藏表格文件")
        #     return

        st.info("开始转换...")
        results = convert_files(source_file, hidden_file)

        if results:
            # 显示结果信息
            st.success("程序执行成功！")
            st.info(f"订单录入文件：{results['order']['filename']}，共 {results['order']['count']} 条记录")
            st.info(f"工件导入文件：{results['workpiece']['filename']}，共 {results['workpiece']['count']} 条记录")
            
            # 提供下载按钮
            st.download_button(
                label=f"下载 {results['order']['filename']}",
                data=results['order']['buffer'],
                file_name=results['order']['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.download_button(
                label=f"下载 {results['workpiece']['filename']}",
                data=results['workpiece']['buffer'],
                file_name=results['workpiece']['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("程序执行失败！请检查错误信息")


if __name__ == "__main__":
    main()