import pandas as pd
import os
import sys
from datetime import datetime
import traceback
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
from openpyxl.styles import NamedStyle


def print_banner():
    """打印程序标题"""
    print("=" * 60)
    print("           益模订单转换工具 v1.0")
    print("        非开发人员专用版本")
    print("=" * 60)
    print()


def select_source_file():
    """选择何氏订单总表文件"""
    print("📁 请选择何氏订单总表文件（Excel格式）...")

    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    file_path = filedialog.askopenfilename(
        title="选择何氏订单总表文件",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
    )

    root.destroy()

    if not file_path:
        print("❌ 未选择文件，程序退出")
        return None

    print(f"✅ 已选择文件: {os.path.basename(file_path)}")
    print()
    return file_path

from copy import copy
from openpyxl.styles import NamedStyle

def copy_sheet(source_wb, source_sheet_name, target_wb, new_sheet_name=None):
    """复制工作表（包含完整格式）"""
    source_sheet = source_wb[source_sheet_name]
    new_name = new_sheet_name or source_sheet_name
    target_sheet = target_wb.create_sheet(new_name)

    # 复制单元格内容和样式
    for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row,
                                     min_col=1, max_col=source_sheet.max_column):
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            
    # 复制列宽
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, source_sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    
    # 复制行高
    for row in range(1, source_sheet.max_row + 1):
        if row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    # 复制合并单元格
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merged_cells.add(str(merged_range))

    # 复制工作表属性
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.conditional_formatting = copy(source_sheet.conditional_formatting)

    # 修正：正确复制命名样式
    for style in source_wb.named_styles:
        target_style_names = []
        for s in target_wb.named_styles:
            if hasattr(s, 'name'):
                target_style_names.append(s.name)
            elif isinstance(s, str):
                target_style_names.append(s)
                # 然后检查当前源样式名称是否已存在于目标工作簿的样式名称列表中
    for style in source_wb.named_styles:
        # 同样检查源样式是否为有效对象
        if hasattr(style, 'name'):
            style_name = style.name
        elif isinstance(style, str):
            style_name = style
        else:
            continue  # 跳过无效的样式对象
            
        if style_name not in target_style_names:
            new_style = NamedStyle(name=style_name)
            if hasattr(style, 'font'):
                new_style.font = copy(style.font)
            if hasattr(style, 'border'):
                new_style.border = copy(style.border)
            if hasattr(style, 'fill'):
                new_style.fill = copy(style.fill)
            if hasattr(style, 'number_format'):
                new_style.number_format = copy(style.number_format)
            if hasattr(style, 'protection'):
                new_style.protection = copy(style.protection)
            if hasattr(style, 'alignment'):
                new_style.alignment = copy(style.alignment)
            target_wb.add_named_style(new_style)

    return target_sheet

def convert_files(source_file):
    """执行文件转换"""
    try:
        print("📖 正在读取何氏订单总表...")
        df_source = pd.read_excel(source_file)
        print(f"✅ 源文件读取成功，共 {len(df_source)} 行数据")

        # ==================== 生成订单录入文件 ====================
        print("\n🔄 正在生成订单录入文件...")

        # 按生产单号去重，只保留第一行数据
        df_unique = df_source.drop_duplicates(subset=['生产单号'], keep='first')
        print(f"✅ 按生产单号去重完成，共 {len(df_unique)} 条记录")

        # 创建订单录入数据
        order_data = []
        for index, row in df_unique.iterrows():
            new_row = {
                '项目名称': str(row['制品名称']),
                '项目编号': str(row['制品名称']),
                '项目预估交货期': row['下单日期'].strftime('%Y-%m-%d'),
                '模具名称': str(row['制品名称']),
                '模具编号': str(row['生产单号']),
                '预估交货期': row['交期'].strftime('%Y-%m-%d'),
                '模具类型': str(row['类型']),
                '模具阶段': str(row['Unnamed: 7']),
                '数量': 1
            }
            order_data.append(new_row)

        df_order_result = pd.DataFrame(order_data)

        # ==================== 生成工件导入文件 ====================
        print("🔄 正在生成工件导入文件...")

        # 创建工件导入数据（保留所有行，不去重）
        workpiece_data = []
        for index, row in df_source.iterrows():
            # 添加原始工件记录
            base_row = {
                '生产任务号': str(row['生产单号']) + '_T0',
                '件号': str(row['制品名称']) + str(row['部件名称']),
                '工件编码': str(row['制品名称']),
                '工件名称': str(row['部件名称']),
                '数量': int(row['数量']),
                '备注': '',
                '生产单号': str(row['生产单号'])
            }
            workpiece_data.append(base_row)

            # 处理母型合金列
            if pd.notna(row.get('母型合金')) and str(row['母型合金']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '母型合金',
                    '工件编码': '母型合金',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理母型合金板列
            if pd.notna(row.get('母型合金板')) and str(row['母型合金板']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '母型合金板',
                    '工件编码': '母型合金板',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理母型套中套列
            if pd.notna(row.get('母型套中套')) and str(row['母型套中套']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '母型套中套',
                    '工件编码': '母型套中套',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理合金针列
            if pd.notna(row.get('合金针')) and str(row['合金针']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': '合金针',
                    '工件编码': '合金针',
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

            # 处理底座列
            if pd.notna(row.get('底座')) and str(row['底座']).strip():
                workpiece_data.append({
                    '生产任务号': str(row['生产单号']) + '_T0',
                    '件号': f"{row['部件名称']}底座",
                    '工件编码': f"{row['部件名称']}底座",
                    '工件名称': '其他配件',
                    '数量': int(row['数量']),
                    '备注': '',
                    '生产单号': str(row['生产单号'])
                })

        df_workpiece_result = pd.DataFrame(workpiece_data)
        print(f"✅ 工件导入数据生成完成，共 {len(df_workpiece_result)} 条记录")

        # ==================== 选择保存位置 ====================
        print("\n💾 请选择保存结果文件的位置...")

        # 选择保存目录
        root = tk.Tk()
        root.withdraw()
        save_dir = filedialog.askdirectory(title="选择保存结果文件的目录")
        root.destroy()

        if not save_dir:
            print("❌ 未选择保存目录，程序退出")
            return False

        print(f"✅ 已选择保存目录: {save_dir}")

        # ==================== 保存文件 ====================
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 加载隐藏表格文件
        hidden_wb = load_workbook('隐藏表格.xlsx', data_only=True)

        # 保存订单录入文件
        order_filename = os.path.join(save_dir, f'订单录入结果_{timestamp}.xlsx')
        print(f"\n💾 正在保存订单录入结果到 {os.path.basename(order_filename)}...")
        
        # 初始化工作簿并确保没有默认空表
        order_wb = Workbook()
        # 删除默认创建的Sheet
        if 'Sheet' in order_wb.sheetnames:
            del order_wb['Sheet']
       
        # 添加订单录入工作表并写入数据
        order_ws = order_wb.create_sheet('订单录入')
        for r in dataframe_to_rows(df_order_result, index=False, header=True):
            order_ws.append(r)
         
        # 复制page工作表
        copy_sheet(hidden_wb, 'page', order_wb, new_sheet_name='page')
        # 设置订单录入表的列宽和行高
        column_widths = {
            'A': 35, 'B': 35, 'C': 15, 'D': 35, 'E': 12,
            'F': 15, 'G': 20, 'H': 12, 'I': 8
        }
        for col_letter, width in column_widths.items():
            order_ws.column_dimensions[col_letter].width = width
        order_ws.row_dimensions[1].height = 25
        for row_num in range(2, len(df_order_result) + 2):
            order_ws.row_dimensions[row_num].height = 20
        
        # 隐藏page工作表
        order_wb['page'].sheet_state = 'hidden'
        
        # 保存订单文件
        order_wb.save(order_filename)
        print(f"✅ 订单录入文件保存完成")


        # 保存工件导入文件
        workpiece_filename = os.path.join(save_dir, f'工件导入结果_{timestamp}.xlsx')
        print(f"💾 正在保存工件导入结果到 {os.path.basename(workpiece_filename)}...")

        # 初始化工作簿并确保没有默认空表
        workpiece_wb = Workbook()
        if 'Sheet' in workpiece_wb.sheetnames:
            del workpiece_wb['Sheet']
        
        # 添加工件信息工作表并写入数据
        workpiece_ws = workpiece_wb.create_sheet('工件信息')
        for r in dataframe_to_rows(df_workpiece_result, index=False, header=True):
            workpiece_ws.append(r)
        # 复制page工作表
        copy_sheet(hidden_wb, 'page2', workpiece_wb, new_sheet_name='page')
        # 设置工件信息表的列宽和行高
        workpiece_column_widths = {
            'A': 15, 'B': 50, 'C': 35, 'D': 20, 'E': 8, 'F': 10, 'G': 12
        }
        for col_letter, width in workpiece_column_widths.items():
            workpiece_ws.column_dimensions[col_letter].width = width
        workpiece_ws.row_dimensions[1].height = 25
        for row_num in range(2, len(df_workpiece_result) + 2):
            workpiece_ws.row_dimensions[row_num].height = 20
        
        # 隐藏page工作表
        workpiece_wb['page'].sheet_state = 'hidden'
        
        # 保存工件文件
        workpiece_wb.save(workpiece_filename)
        print(f"✅ 工件导入文件保存完成")

        # ==================== 输出结果统计 ====================
        print(f"\n🎉 转换完成！")
        print(f"📊 订单录入文件：{os.path.basename(order_filename)}，共 {len(df_order_result)} 条记录")
        print(f"📊 工件导入文件：{os.path.basename(workpiece_filename)}，共 {len(df_workpiece_result)} 条记录")
        print(f"📁 文件保存在：{save_dir}")

        return True

    except Exception as e:
        print(f"\n❌ 转换过程中出现错误:")
        print(f"错误信息: {str(e)}")
        print("\n详细错误信息:")
        traceback.print_exc()
        return False


def main():
    """主函数"""
    try:
        print_banner()

        print("📋 程序说明：")
        print("本工具将何氏订单总表的数据转换为两个文件：")
        print("1. 订单录入结果 - 按生产单号去重后的模具级别信息")
        print("2. 工件导入结果 - 包含所有工件详细信息")
        print()

        # 选择源文件
        source_file = select_source_file()
        if not source_file:
            return

        print("🚀 开始转换...")
        print()

        # 执行转换
        success = convert_files(source_file)

        if success:
            print("\n✅ 程序执行成功！")
            print("📁 生成的文件在您选择的目录中")
        else:
            print("\n❌ 程序执行失败！")
            print("请检查错误信息并联系技术支持")

        print()
        input("按回车键退出...")

    except KeyboardInterrupt:
        print("\n\n⚠️ 程序被用户中断")
        input("按回车键退出...")
    except Exception as e:
        print(f"\n❌ 程序出现未知错误: {str(e)}")
        print("请联系技术支持")
        input("按回车键退出...")


if __name__ == "__main__":
    main()